import time
import requests
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook

class Aims:
    def __init__(self, log_request, log_data, log_return):
        self.login_request = log_request
        self.login_data = log_data
        self.login_return = log_return

    def login(self, login_url, username, b_month, b_day, b_year, password):
        self.login_data["txtUser"] = username
        self.login_data["cboMonth"] = b_month
        self.login_data["cboDay"] = b_day
        self.login_data["cboYear"] = b_year
        self.login_data["txtPwd"] = password
        try:
            login_response = self.login_request.post(login_url, self.login_data).text
            if self.login_return["authentication"] in login_response:
                print("[ ✘ ] Authentication failed! Check your username or password!")
                return False
            elif self.login_return["attempts"] in login_response:
                print("[ ✘ ] Too many account access attempts.")
                return False
            elif self.login_return["form_error"] in login_response.split("\n")[29]:
                print("[ ✘ ] Authentication failed! Something Went Wrong...")
                return False
            else:
                print("[ ✔ ] Account Verification Successful ...")
                return True
        except Exception as error:
            print(error)
            time.sleep(3)
            sys.exit(0)

    def yield_grades(self, grades_url):
        try:
            grades_response = self.login_request.get(grades_url).text
            cookies = str(self.login_request.cookies).split("=")[1].split(" ")[0]
            print("HTTP Cookie: %s" % cookies)
            return grades_response
        except Exception as error:
            print(error)
            time.sleep(3)
            sys.exit(0)

class Serialize:
    def __init__(self, grades):
        self.grades_list = grades

    def yield_student_info(self, index):
        stud_info = {"student_id": [], "fullname": []}
        stud_info["student_id"].append(self.grades_list[index].split("<b>")[1].split("</b>")[0].split("(")[1].split(")")[0])
        stud_info["fullname"].append(self.grades_list[index].split("<b>")[1].split("</b>")[0].split("(")[0])
        return stud_info

    def yield_school_status(self, index, digit):
        return self.grades_list[sum([index, digit])].split("<b>")[1].split("</b>")[0]

    def yield_status_name(self, index, digit):
        return self.grades_list[sum([index, digit])].split("<u>")[1].split("</u>")[0]

    def yield_terms_fe_abs(self, index, digit):
        return self.grades_list[sum([index, digit])][91:].split("<font color='000000'>")[1:]

    def yield_ave_gs(self, index, digit):
        return self.grades_list[sum([index, digit])][91:].split("<font color='000000'>")[1:][grades_list[sum([index, digit])].count("<font color='000000'>") - 1].split("<font color='#000000'>")[1:]

    def yield_table_data_index(self, number_index):
        return self.grades_list[number_index].split("</td>")[0]

    def yield_grade_values(self, index, digit):
        return self.grades_list[sum([index, digit])].split("<td>")[1].split("</td>")[0]

class PyExcel:
    def __init__(self, school_status_name, school_status, columns, rows):
        self.school_status_name = school_status_name
        self.school_status = school_status
        self.table_columns = columns
        self.table_rows = rows

    def append_grade_values(self, filename, sheet_title):
        if not os.path.exists(filename):
            work_book = Workbook()
            work_sheet = work_book.active
            work_sheet.title = sheet_title
        else:
            work_book = load_workbook(filename)
            work_book.create_sheet(sheet_title)
            work_book.active = len(work_book.sheetnames) - 1
            work_sheet = work_book.active
            work_sheet.title = sheet_title

        ws_stat_name_index = ["A1", "C1", "A2", "C2", "A3", "C3", "A4"]
        ws_status_index = ["B1", "D1", "B2", "D2", "B3", "D3", "B4"]
        counter_index = 0
        for each_status_name in self.school_status_name:
            work_sheet[ws_stat_name_index[counter_index]] = each_status_name
            work_sheet[ws_stat_name_index[counter_index]].font = Font(bold=True)
            counter_index = counter_index + 1
        counter_index = 0
        for each_status in self.school_status:
            work_sheet[ws_status_index[counter_index]] = each_status
            work_sheet[ws_status_index[counter_index]].font = Font(bold=False)
            counter_index = counter_index + 1
        col_letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        work_sheet.append([])
        work_sheet.append(self.table_columns)
        for cols in range(0, len(self.table_columns)):
            work_sheet["%s%s" % (col_letters[cols], 6)].font = Font(bold=True)
        for each_rows in self.table_rows:
            work_sheet.append(each_rows)

        work_book.save(filename)

class Arguments:
    def __init__(self, sys_params):
        self.system_parameters = sys_params
        self.used_params = []

    def yield_parameters(self):
        return self.used_params

    def check_arguments(self, parameters):
        try:
            if len(parameters) != 0:
                if parameters[0] in self.system_parameters[-5:]:
                    return None
                else:
                    for each_params in parameters:
                        if each_params in self.system_parameters[:6]:
                            self.used_params.append(each_params)
                    if len(self.used_params) == 0:
                        return False
            else:
                return False
        except IndexError as error:
            return error

        return True

aims_urls = {
    "login_url": "https://upangsms.phinma.edu.ph/upang/process/validate.php?userType=1",
    "grades_url": "https://upangsms.phinma.edu.ph/upang/students/grades.php?mainID=106&menuDesc=Grades"
    }
login_return = {
    "authentication": "Authentication failed! Check your username or password!",
    "attempts": "Too many login attempts.",
    "form_error": "doSubmit()"
    }
login_data = {
    "txtUser": "",
    "cboMonth": "",
    "cboDay": "",
    "cboYear": "",
    "txtPwd": ""
    }

if __name__ == "__main__":
    if not os.path.isdir("Accounts"):
        os.mkdir("Accounts")
    session = requests.session()
    account = Aims(session, login_data, login_return)
    arguments = Arguments(["-U", "-u", "-B", "-b", "-P", "-p", "--H", "--h", "-h", "-H", "--help"])
    params_check = arguments.check_arguments(sys.argv[1:])
    if params_check is not True:
        if params_check is not None:
            print("Invalid Arguments. Please enter an appropriate one")
            time.sleep(2)
            sys.exit(0)
    credentials = {"student_id": None, "b_month": None, "b_day": None, "b_year": None, "password": None}
    for i in range(0, len(sys.argv[1:])):
        try:
            if sys.argv[i] == "-U" or sys.argv[i] == "-u":
                credentials["student_id"] = sys.argv[sum([i, 1])]
            elif sys.argv[i] == "-B" or sys.argv[i] == "-b":
                credentials["b_month"] = (sys.argv[sum([i, 1])]).split("|")[0]
                credentials["b_day"] = (sys.argv[sum([i, 1])]).split("|")[1]
                credentials["b_year"] = (sys.argv[sum([i, 1])]).split("|")[2]
            elif sys.argv[i] == "-P" or sys.argv[i] == "-p":
                credentials["password"] = (sys.argv[sum([i, 1])])
        except IndexError:
            print("Invalid Arguments. Please enter an appropriate one")
            time.sleep(2)
            sys.exit(0)
    print("Authenticating account ...", end="\r", flush=True)
    if account.login(aims_urls["login_url"], credentials["student_id"], credentials["b_month"], credentials["b_day"], credentials["b_year"], credentials["password"]):
        grades_list = list(account.yield_grades(aims_urls["grades_url"]).split("\n"))
        account_info = Serialize(grades_list)
        student_id = account_info.yield_student_info(117)["student_id"][0]
        fullname = account_info.yield_student_info(117)["fullname"][0]
        status = {"index": 139, "digits": 173, "columns": [], "rows": [], "temp_rows": []}
        print("Account: %s ( %s )" % (fullname, student_id))
        collect_stats = "Collecting Grades From Grading Portal, Please Wait "
        status_bar_count = 1
        time.sleep(1)
        while True:
            print("%s%s" % (collect_stats, "█ " * status_bar_count), end="\r", flush=True)
            try:
                sy_name = account_info.yield_school_status(status["index"], 0)
                school_year = account_info.yield_status_name(status["index"], 1)
                sem_name = account_info.yield_school_status(status["index"], 2)
                semester = account_info.yield_status_name(status["index"], 3)
                admission_name = account_info.yield_school_status(status["index"], 6)
                admission_status = account_info.yield_status_name(status["index"], 7)
                scholastic_name = account_info.yield_school_status(status["index"], 8)
                scholastic_status = account_info.yield_status_name(status["index"], 9)
                code_name = account_info.yield_school_status(status["index"], 12)
                course_code = account_info.yield_status_name(status["index"], 13)
                c_description = account_info.yield_school_status(status["index"], 14)
                course_description = account_info.yield_status_name(status["index"], 15)
                gpa_name = account_info.yield_school_status(status["index"], 18).split("(")[0]
                gpa = account_info.yield_status_name(status["index"], 19)

                s_code_name = account_info.yield_school_status(status["index"], 27)
                desc_name = account_info.yield_school_status(status["index"], 28)
                fac_name = account_info.yield_school_status(status["index"], 29)
                units_name = account_info.yield_school_status(status["index"], 30)
                section_name = account_info.yield_school_status(status["index"], 31)

                status["columns"].append(s_code_name)
                status["columns"].append(desc_name)
                status["columns"].append(fac_name)
                status["columns"].append(units_name)
                status["columns"].append(section_name)

                terms_fe_abs = account_info.yield_terms_fe_abs(status["index"], 31)
                for each_col in terms_fe_abs[:-1]:
                    status["columns"].append(each_col.split("<b>")[1].split("</b>")[0])
                else:
                    status["columns"].append(terms_fe_abs[-1:][0].split("<font color='#000000'>")[0].split("<b>")[1].split("</b>")[0])
                ave_gs = account_info.yield_ave_gs(status["index"], 31)
                status["columns"].append(ave_gs[0].split("<b>")[1].split("</b>")[0])
                status["columns"].append(ave_gs[1].split("<b>")[1].split("</b>")[0])

                while True:
                    status["temp_rows"] = []
                    if "<td>" in account_info.yield_table_data_index(status["digits"]):
                        subject_code = account_info.yield_grade_values(status["digits"], 1)
                        description = account_info.yield_grade_values(status["digits"], 2)
                        faculty = account_info.yield_grade_values(status["digits"], 3)
                        units = account_info.yield_grade_values(status["digits"], 4)
                        section = account_info.yield_grade_values(status["digits"], 5)
                        status["temp_rows"].append(subject_code)
                        status["temp_rows"].append(description)
                        status["temp_rows"].append(faculty)
                        status["temp_rows"].append(units)
                        status["temp_rows"].append(section)
                        grades_abs_fe = grades_list[sum([status["digits"], 5])].split("<td>")
                        for each_gaf in grades_abs_fe[2:]:
                            status["temp_rows"].append(each_gaf.split("</td>")[0])
                        status["rows"].append(status["temp_rows"])
                    else:
                        break
                    status["digits"] += 8

                status["index"] = status["digits"] + 4
                status["digits"] += 38

                py_excel_sheet = PyExcel([sy_name, sem_name, admission_name, scholastic_name, code_name, c_description, gpa_name],
                                         [school_year, semester, admission_status, scholastic_status, course_code, course_description, gpa],
                                         status["columns"], status["rows"])
                py_excel_sheet.append_grade_values("aims.grades (%s).xlsx" % student_id, "%s-%s" % (school_year, semester))

                status["columns"] = []
                status["rows"] = []

                if status_bar_count > 10:
                    status_bar_count = 1
                status_bar_count += 1

            except (ValueError, IndexError) as sys_error:
                break

        time.sleep(1)
        print("[ ✔ ] Grades Collected Successfully... \t\t\t\t\t\t\t")
        time.sleep(1)
        print(r"Check here: %s\aims.grades (%s).xlsx" % (os.getcwd(), student_id))
        os.startfile("aims.grades (%s).xlsx" % student_id)

    else:
        print("[ ✘ ] Failed To Login Account. Try Again!")
        time.sleep(3)
        sys.exit(0)
